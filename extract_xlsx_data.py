"""
从 Excel 提取图片和原因文本，写入 data/xlsx_records.json
- 图片解压到 data/ 目录（文件名保持 Excel 内部名称，如 image1.png）
- 跳过非图片格式（.emf 等）
- 无图片的行也记录，image_path 为 null
"""
import json
import re
import shutil
import zipfile
from pathlib import Path

import openpyxl

XLSX_PATH = Path(__file__).parent / "data" / "报修工单图片及报修问题描述.xlsx"
DATA_DIR = Path(__file__).parent / "data"
OUTPUT_JSON = DATA_DIR / "xlsx_records.json"

IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp", ".tiff"}


def extract():
    # 1. B 列原因文本
    wb = openpyxl.load_workbook(str(XLSX_PATH))
    ws = wb.active
    b_col: dict[int, str] = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        cell_b = row[1]
        if cell_b.value:
            b_col[cell_b.row] = str(cell_b.value).strip()

    # 2. C 列图片映射：row -> 内部文件名
    with zipfile.ZipFile(str(XLSX_PATH)) as z:
        sheet_xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8")
        rels_xml = z.read("xl/richData/_rels/richValueRel.xml.rels").decode("utf-8")

        rid_to_file = dict(
            re.findall(r'Id="(rId\d+)"[^>]*Target="\.\./media/([^"]+)"', rels_xml)
        )
        c_cells = re.findall(r'<c r="C(\d+)"[^>]*vm="(\d+)"', sheet_xml)
        row_to_img: dict[int, str] = {
            int(r): rid_to_file.get(f"rId{vm}", "")
            for r, vm in c_cells
        }

        # 3. 解压图片到 data/
        extracted = 0
        skipped = 0
        for img_name in rid_to_file.values():
            ext = Path(img_name).suffix.lower()
            if ext not in IMAGE_EXTS:
                print(f"  [SKIP non-image] {img_name}")
                skipped += 1
                continue
            src = f"xl/media/{img_name}"
            dst = DATA_DIR / img_name
            if dst.exists():
                continue  # already extracted
            with z.open(src) as src_f, open(dst, "wb") as dst_f:
                shutil.copyfileobj(src_f, dst_f)
            extracted += 1

    print(f"图片解压完成：新增 {extracted} 张，跳过非图片 {skipped} 个")

    # 4. 合并记录
    all_rows = sorted(set(b_col) | set(row_to_img))
    records = []
    for row_num in all_rows:
        reason = b_col.get(row_num, "")
        img_name = row_to_img.get(row_num, "")

        # 跳过非图片格式（如 .emf）
        if img_name and Path(img_name).suffix.lower() not in IMAGE_EXTS:
            print(f"  [row={row_num}] 跳过非图片文件: {img_name}")
            continue

        img_path = str(DATA_DIR / img_name) if img_name else None

        records.append({
            "row_number": row_num,
            "original_reason": reason,
            "image_filename": img_name or None,
            "image_path": img_path,
        })

    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)

    has_img = sum(1 for r in records if r["image_filename"])
    no_img = sum(1 for r in records if not r["image_filename"])
    print(f"共 {len(records)} 条记录：{has_img} 有图片，{no_img} 无图片")
    print(f"结果写入 {OUTPUT_JSON}")


if __name__ == "__main__":
    extract()
